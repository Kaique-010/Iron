from datetime import date, timedelta
import json
import os
from django.conf import settings
from io import BytesIO
from django.db import IntegrityError, transaction
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
import openpyxl
from django.views.generic.edit import FormView
from django.forms import modelformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, UpdateView, DetailView, DeleteView
from django.forms import inlineformset_factory
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseBadRequest
from django.db.models import Sum, Max
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.db import connection, connections
from collections import namedtuple
from Financeiro.models import FormasRecebimento
from Pessoas.models import Pessoas
from .models import Pedido, ItemPedido
from .forms import CRMForm, ItemPedidoForm, PedidoForm
from Produtos.models import Produtos
from Financeiro.forms import DateRangeForm
from django.core.mail import send_mail
from django.core.exceptions import MultipleObjectsReturned
from django.db.models import Count


def set_empresa_database(empresa):
    db_alias = empresa.database  
    if db_alias not in connections:
        raise IntegrityError(f"Banco de dados {db_alias} não configurado.")
    connections['default'] = connections[db_alias]


def dictfetchall(cursor):
    """Converte todas as linhas do cursor em um dicionário."""
    columns = [col[0] for col in cursor.description]
    Row = namedtuple('Row', columns)
    return [Row(*row)._asdict() for row in cursor.fetchall()]

# Formset para ItemPedido
ItemPedidoFormSet = inlineformset_factory(Pedido, ItemPedido, form=ItemPedidoForm, extra=1, can_delete=True)

class PedidosListView(ListView):
    model = Pedido
    template_name = 'pedidoslistas.html'
    context_object_name = 'pedidos'
    paginate_by = 5

    def get_queryset(self):
        if not self.request.user.empresa:
            raise IntegrityError("Usuário não associado a uma empresa")
        
        set_empresa_database(self.request.user.empresa)  
        
        licenca_id = self.request.session.get('licenca_id')  
        queryset = super().get_queryset().using(licenca_id)
        
        pedido = self.request.GET.get('pedido')
        cliente = self.request.GET.get('cliente')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if pedido:
            queryset = queryset.filter(id__icontains=pedido)
        if cliente:
            queryset = queryset.filter(cliente__nome__icontains=cliente)
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                queryset = queryset.filter(data__gte=start_date)
        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                queryset = queryset.filter(data__lte=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = DateRangeForm(self.request.GET or None)

        start_date = self.request.GET.get('start_date', '')
        end_date = self.request.GET.get('end_date', '')

        context['form'] = form
        context['total_count'] = self.get_queryset().count()
        context['total_value'] = self.get_queryset().aggregate(total_value=Sum('total'))['total_value'] or 0
        context['start_date'] = start_date
        context['end_date'] = end_date

        return context
    

class PedidoCreateView(FormView):
    template_name = 'pedidoscriar.html' 
    form_class = PedidoForm  

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Cria o formset para o ItemPedido
        ItemPedidoFormSet = modelformset_factory(ItemPedido, form=ItemPedidoForm, extra=1)
      
        if self.request.POST:
            context['formset'] = ItemPedidoFormSet(self.request.POST)
        else:
            context['formset'] = ItemPedidoFormSet(queryset=ItemPedido.objects.none())
        
        return context

    def form_valid(self, form):
        # Salva o pedido (associando a empresa do usuário)
        pedido = form.save(commit=False)
        pedido.empresa = self.request.user.empresa  # Atribui a empresa do usuário ao pedido
        pedido.save()  # Agora o pedido é salvo no banco e tem um PK

        # Obtém o formset de ItemPedido
        formset = modelformset_factory(ItemPedido, form=ItemPedidoForm, extra=1)(self.request.POST)

        if formset.is_valid():
            # Associa cada item ao pedido
            for item_form in formset:
                item = item_form.save(commit=False)
                item.pedido = pedido  # Associa o item ao pedido
                item.save()  # Salva o item

            # Redireciona para a página de sucesso
            return redirect('success')  # Você pode mudar 'success' para o nome de URL desejado
        else:
            # Caso o formset seja inválido, renderiza a página com o formset e o formulário principal
            return self.render_to_response(self.get_context_data(form=form, formset=formset))

    def form_invalid(self, form):
        # Caso o formulário principal ou o formset sejam inválidos
        formset = modelformset_factory(ItemPedido, form=ItemPedidoForm, extra=1)(self.request.POST)
        return self.render_to_response(self.get_context_data(form=form, formset=formset))

def buscar_produtos(request):
    query = request.GET.get('q', '')
    produtos = Produtos.objects.filter(nome__icontains=query)
    resultado = [{'id': produto.id, 'nome': produto.nome} for produto in produtos]
    return JsonResponse(resultado, safe=False)

class PedidosUpdateView(UpdateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedidoseditar.html'
    success_url = reverse_lazy('pedidoslistas')

    def get_form(self, form_class=None):
        if not self.request.user.empresa:
            raise IntegrityError("Usuário não associado a uma empresa")
    
        set_empresa_database(self.request.user.empresa)
   
        return super().get_form(form_class)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            formset = ItemPedidoFormSet(self.request.POST, instance=self.object)
        else:
            formset = ItemPedidoFormSet(instance=self.object)
        context['formset'] = formset
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Atualiza o formset com a instância do pedido existente ou recém-atualizado
        formset = ItemPedidoFormSet(self.request.POST, instance=self.object)
        if formset.is_valid():
            formset.save()  # Salva o formset com os itens
        else:
            return self.render_to_response(self.get_context_data(form=form, formset=formset))
        
        return response
    
class PedidosDetailView(DetailView):
    model = Pedido
    template_name = 'pedidosdetalhe.html'

    def get_object(self):
        if not self.request.user.empresa:
            raise IntegrityError("Usuário não associado a uma empresa")
        
        set_empresa_database(self.request.user.empresa)
        return super().get_object()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido = self.object

        # Calcula os totais
        total_quantidade = sum(item.quantidade for item in pedido.itens.all())
        total_valor = sum(item.total for item in pedido.itens.all())  # Supondo que item.total seja o subtotal do item

        # Adiciona os totais ao contexto
        context['total_quantidade'] = total_quantidade
        context['total_valor'] = total_valor
        return context
class PedidosDeleteView(DeleteView):
    model = Pedido
    template_name = 'pedidosexcluir.html'
    success_url = reverse_lazy('pedidoslistas')
    
    def get_object(self):
        if not self.request.user.empresa:
            raise IntegrityError("Usuário não associado a uma empresa")
        
        set_empresa_database(self.request.user.empresa)  # Configura o banco de dados
        return super().get_object()
    

def pedido_itens(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    context = {'pedido': pedido}
    return render(request, 'pedido_itens.html', context)


def criar_planilha_pedidos(pedidos):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos'
    columns = ['Nº Pedido', 'Data', 'Cliente', 'Total Pedido', 'Vendedor']
    worksheet.append(columns)
    for pedido in pedidos:
        worksheet.append([
            pedido.id,
            pedido.data.strftime('%d-%m-%y'),
            pedido.cliente.nome if pedido.cliente else '',
            pedido.total,
            pedido.nome_vendedor
        ])
    return workbook

def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.all()
    workbook = criar_planilha_pedidos(pedidos)
    worksheet_itens = workbook.create_sheet(title='Itens dos Pedidos')
    columns_itens = ['Nº Pedido', 'Item', 'Quantidade', 'Preço Unitário', 'Subtotal']
    worksheet_itens.append(columns_itens)
    itens_pedidos = ItemPedido.objects.select_related('pedido').all()
    for item in itens_pedidos:
        worksheet_itens.append([
            item.pedido.id,
            item.produto.nome,
            item.quantidade,
            item.preco_unitario,
            item.quantidade * item.preco_unitario,
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    return response

def crm(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    cliente = request.GET.get('cliente')

    pedidos = Pedido.objects.all()

    if start_date and end_date:
        pedidos = pedidos.filter(data__range=[start_date, end_date])

    if cliente:
        pedidos = pedidos.filter(nome_cliente__icontains=cliente)

    # Subquery para pegar o último pedido de cada cliente
    subquery = Pedido.objects.values('cliente_id').annotate(max_id=Max('id'))
    pedidos = pedidos.filter(id__in=[s['max_id'] for s in subquery])

    context = {
        'crm_data': pedidos,
        'start_date': start_date,
        'end_date': end_date,
        'form': CRMForm(initial={'start_date': start_date, 'end_date': end_date}),
    }
    return render(request, 'crm.html', context)

def marcar_contato_realizado(request, pedido_id):
    try:
        pedido = Pedido.objects.get(id=pedido_id)
        pedido.contato_realizado = True
        pedido.data_contato = now().date()
        pedido.save()
    except Pedido.DoesNotExist:
        return redirect('crm')
    except MultipleObjectsReturned:
        return redirect('crm')

    return redirect('detalhar_contato', pedido_id=pedido_id)



def detalhar_contato(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    context = {'pedido': pedido}
    return render(request, 'detalhar_contato.html', context)



def exportar_crm_excel(request):
    pedidos = Pedido.objects.filter(contato_realizado=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CRM Export'
    columns = ['Cliente', 'Data do Contato', 'Nº Pedido', 'Total']
    worksheet.append(columns)
    for pedido in pedidos:
        worksheet.append([
            pedido.cliente.nome if pedido.cliente else '',
            pedido.data_contato,
            pedido.id,
            pedido.total
        ])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=crm_export.xlsx'
    return response


def dashboard(request):
    # Filtra os pedidos com base na licença do usuário
    licenca_id = request.session.get('licenca_id')
    pedidos = Pedido.objects.using(licenca_id).all()
    
    # Obtém as métricas necessárias
    total_pedidos = pedidos.count()
    total_vendas = pedidos.aggregate(total=Sum('total'))['total'] or 0
    pedidos_por_cliente = pedidos.values('cliente__nome').annotate(total=Count('id'))
    
    # Organiza os dados para gráficos ou relatórios
    cliente_dados = {cliente['cliente__nome']: cliente['total'] for cliente in pedidos_por_cliente}
    
    context = {
        'total_pedidos': total_pedidos,
        'total_vendas': total_vendas,
        'clientes': cliente_dados.items(),
    }
    
    return render(request, 'dashboard.html', context)


def enviar_emails_inativos(request):
    hoje = date.today()
    
    if request.method == 'GET':
        dias_para_contato = int(request.GET.get('dias', 180))
        data_limite = hoje - timedelta(days=dias_para_contato)
        
        query = """
            SELECT 
                p.cliente_id AS cliente_id,
                pe.email AS email,
                MAX(p.data) AS ultima_compra,
                MAX(p.total) AS total,
                p.nome_cliente AS nome_cliente
            FROM pedidos p
            JOIN pessoas pe ON p.cliente_id = pe.id
            WHERE p.data <= %s
            GROUP BY p.cliente_id, pe.email, p.nome_cliente
                    """
        
        with connection.cursor() as cursor:
            cursor.execute(query, [data_limite])
            rows = cursor.fetchall()
            clientes_inativos = [
                {
                    'cliente_id': row[0],
                    'email': row[1],
                    'ultima_compra': row[2],
                    'total': row[3],
                    'nome_cliente': row[4],
                }
                for row in rows
            ]
        
        context = {
            'clientes_inativos': clientes_inativos,
            'dias_para_contato': dias_para_contato,
        }
        return render(request, 'emails_inativos.html', context)
    
    elif request.method == 'POST':
        clientes_ids = request.POST.getlist('clientes')
        mensagem = request.POST.get('mensagem')
        
        if not clientes_ids or not mensagem:
            return HttpResponseBadRequest("Nenhum cliente selecionado ou mensagem vazia.")
        
        query = """
            SELECT email FROM pessoas
            WHERE id IN %s
        """
        
        with connection.cursor() as cursor:
            cursor.execute(query, [tuple(clientes_ids)])
            emails = [row[0] for row in cursor.fetchall()]
        
        send_mail(
            subject='Estamos com saudades!',
            message=mensagem,
            from_email='sistema@empresa.com',
            recipient_list=emails,
        )
        
        return redirect('enviar_emails_inativos')  
    
    return HttpResponseBadRequest("Método inválido.")


def clientes_inativos_por_ultima_compra(request):
    hoje = date.today()
    dias_para_contato = int(request.GET.get('dias', 180))
    data_limite = hoje - timedelta(days=dias_para_contato)

    query = """
        SELECT 
            p.cliente_id,
            pe.email,
            MAX(p.data) AS ultima_compra,
            p.total AS total
        FROM Pedidos_pedido p
        JOIN Pessoas_pessoas pe ON p.cliente_id = pe.id
        GROUP BY p.cliente_id, pe.email
        HAVING MAX(p.data) <= %s
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [data_limite])
        clientes_inativos = cursor.fetchall()

    context = {
        'clientes_inativos': clientes_inativos,
        'dias_para_contato': dias_para_contato,
    }
    return render(request, 'clientes_inativos.html', context)


def emails(request):
    hoje = date.today()
    dias_para_contato = int(request.GET.get('dias', 180))
    data_limite = hoje - timedelta(days=dias_para_contato)

    query = """
        SELECT 
            DISTINCT 
            p.cliente_id,
            pe.email,
            MAX(p.data) AS ultima_compra,
            p.total AS total
        FROM Pedidos_pedido p
        JOIN Pessoas_pessoas pe ON p.cliente_id = pe.id
        WHERE p.data <= %s
        GROUP BY p.cliente_id, pe.email
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [data_limite])
        clientes_inativos = cursor.fetchall()

    context = {
        'clientes_inativos': clientes_inativos,
        'data_limite': data_limite,
        'dias_para_contato': dias_para_contato
    }

    return render(request, 'emails_inativos.html', context)


def gerar_pedido_pdf(request, pedido_id):
    # Busca o pedido no banco de dados
    try:
        pedido = Pedido.objects.get(id=pedido_id)
    except Pedido.DoesNotExist:
        return HttpResponse("Pedido não encontrado.", status=404)

    # Renderiza o HTML com os dados do pedido
    html_content = render_to_string('pedido_pdf.html', {'pedido': pedido})

    # Caminho para salvar o PDF no diretório de media
    pdf_path = os.path.join(settings.MEDIA_ROOT, f'pedidos/pedido_{pedido_id}.pdf')
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)  # Cria o diretório se não existir

    # Gera o PDF
    HTML(string=html_content).write_pdf(target=pdf_path)

    # Retorna o PDF gerado como resposta
    with open(pdf_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="pedido_{pedido_id}.pdf"'
        return response